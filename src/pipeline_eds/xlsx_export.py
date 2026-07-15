# src/pipeline_eds/xlsx_export.py
from __future__ import annotations
import re
import io
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

from .context import config_mngr, APP_DIR, APP_SERVICE
from .helpers import iso_time

logger = logging.getLogger(__name__)

def export_xlsx_for_results(results, idcs, plant_name):

    wb = Workbook()
    ws = wb.active
    ws.title = "Trend Data"

    def extract_timestamps_from_results(results):
        all_timestamps = set()
        for rows in results:
            for row in rows:
                if row.get("ts") is not None:
                    all_timestamps.add(row.get("ts"))
                            
        sorted_timestamps = sorted(list(all_timestamps))
        return sorted_timestamps
    def extract_data_matrix_from_results(results, sorted_timestamps):
        data_matrix = {ts: {} for ts in sorted_timestamps}
        for idx, rows in enumerate(results):
            sensor_id = idcs[idx]
            for row in rows:
                ts = row.get("ts")
                if ts is not None:
                    data_matrix[ts][sensor_id] = row.get("value")
        return data_matrix
    
    sorted_timestamps = extract_timestamps_from_results(results)
    data_matrix = extract_data_matrix_from_results(results, sorted_timestamps)
    
    headers = ["Timestamp"] + idcs
    ws.append(headers)
        
    # Make the header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        
    # all trend data
    for ts in sorted_timestamps:
        row_data = [iso_time(ts)]
        for sensor_id in idcs:
                
            row_data.append(data_matrix[ts].get(sensor_id, ""))
            
        existing_times = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        formatted_time = iso_time(ts)
        
        if formatted_time in existing_times:
            row_idx = existing_times.index(formatted_time) + 2
            for col_idx, sensor_id in enumerate(idcs, start=2):
                    val = data_matrix[ts].get(sensor_id, "")
                    if val != "":
                        ws.cell(row=row_idx, column=col_idx, value=val)
        else:
                ws.append(row_data)                
    # Auto-size each column
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Save
    def assess_download_dir(xlsx_export_dir:str|Path|None=None)->Path:
        #FALLBACK_XLSX_EXPORT_DIR = Path.home() / "Downloads"
        DEFAULT_XLSX_EXPORT_DIR = APP_DIR / "exports"
        if xlsx_export_dir is None:
            config_mngr.set(service=APP_SERVICE,item="xlsx_export_dir",value = str(DEFAULT_XLSX_EXPORT_DIR), overwrite = False)
            xlsx_export_dir = config_mngr.get(service=APP_SERVICE,item="xlsx_export_dir")

        # 5. Resolve path if the config value is not empty/null; otherwise fallback
        if xlsx_export_dir and str(xlsx_export_dir).strip():
            resolved_dir = Path(xlsx_export_dir).expanduser().resolve()
        else:
            resolved_dir = DEFAULT_XLSX_EXPORT_DIR

        resolved_dir.mkdir(parents=True, exist_ok=True)
        logger.debug(f"{resolved_dir=}")
        return resolved_dir

    def generate_xlsx_export_filename(plant_name):
        from datetime import datetime
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{plant_name}_trend_{now_str}.xlsx"
        return filename
    
    downloads_dir = assess_download_dir()
    filename = generate_xlsx_export_filename(plant_name)
    file_path = downloads_dir / filename
    
    logger.debug(f"XLSX export filepath: {file_path}")
    wb.save(file_path)
    
    return file_path, wb

def save_xlsx_worbook_to_filestream(wb:Workbook)->io.BytesIO:
    # Save to memory stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream